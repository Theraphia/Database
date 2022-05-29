import os
import openpyxl


class Table:
    def __init__(self, t_name, db_name):
        self.path = db_name + "\\" + t_name + ".xlsx"
        flag = os.path.exists(self.path)
        if not flag:
            file = openpyxl.Workbook()
            file.create_sheet()
            file.save(self.path)
            print("成功创建新表格")
        else:
            print("表格已存在!")
        return

    def set_column(self, inst):
        """
        设置表格列名与列属性
        :param inst:指令
        :return:
        """
        l = inst.split(',')
        list0 = []
        list1 = []
        for c in l:
            col = c.split(' ')
            list0.append(col[0])    # 数据类型
            list1.append(col[1])    # 列名
        file = openpyxl.Workbook()
        table = file.active
        for i in range(len(list0)):
            table.cell(1, i+1, list0[i])
            table.cell(2, i+1, list1[i])
        file.save(self.path)
        return

    def delete(self):
        os.remove(self.path)
        print("表格已删除")
        return


class Database:
    def __init__(self, name):
        self.path = name
        flag = os.path.exists(name)
        if not flag:
            os.makedirs(name)
            print("成功创建新数据库")
        else:
            print("数据库已存在!")
        return

    def delete(self):
        os.rmdir(self.path)
        print("数据库已删除")
        return


class System:
    def __init__(self):
        print("---数据库管理系统已启动---")
        return

    def run(self):
        while 1:
            print("请输入指令：")
            str = input()
            if str == "help":
                print("1:")
                print("2:")
                print("3:")
                print("quit:退出")
                return
            if str == "quit":
                print("---数据库管理系统已关闭---")
                return
        return


sys = System()
# sys.run()

db = Database("db1")
t = Table("t1", "db1")
# t.set_column("int a,int b,char c")
t.delete()
db.delete()
